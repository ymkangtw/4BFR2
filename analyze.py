import openpyxl
import json
from pathlib import Path

xlsx = Path(__file__).parent / "R02NEW.xlsx"
out = Path(__file__).parent / "analysis.txt"
wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)

lines = []
lines.append(f"Sheets ({len(wb.sheetnames)}):")
for name in wb.sheetnames:
    ws = wb[name]
    lines.append(f"  - {name}: {ws.max_row} rows x {ws.max_column} cols")

lines.append("")
lines.append("=" * 80)

for name in wb.sheetnames:
    ws = wb[name]
    lines.append("")
    lines.append(f"[Sheet] {name}")
    lines.append(f"  dims: {ws.max_row} rows x {ws.max_column} cols")
    rows = list(ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), values_only=True))
    for i, row in enumerate(rows, 1):
        trimmed = [(str(c)[:60] if c is not None else "") for c in row]
        lines.append(f"  R{i}: " + json.dumps(trimmed, ensure_ascii=False))

# Also dump full content of the small sheet
for name in wb.sheetnames:
    ws = wb[name]
    if ws.max_row <= 30:
        lines.append("")
        lines.append(f"[FULL Sheet] {name}")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            trimmed = [(str(c)[:80] if c is not None else "") for c in row]
            lines.append(f"  R{i}: " + json.dumps(trimmed, ensure_ascii=False))

wb.close()
out.write_text("\n".join(lines), encoding="utf-8")
print(f"Wrote {out}")
